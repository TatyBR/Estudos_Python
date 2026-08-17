from openpyxl import Workbook

# Criado um arquivo excel
wb = Workbook()

# Verificando as abas existentes
print(wb.sheetnames)

# Selecionando a aba
sheet = wb["Sheet"]

# Modificando o nome da aba
# sheet.title = "Primeira Planilha"

# inserindo dados
sheet["A1"].value = "Python"
sheet["B1"].value = "Formação Expert"
sheet["C1"].value = "Automação de Processos"

# inserindo utilizando as coordenadas
# Lembrando que a coluna A seria 1, B 2, C 3.... Inicia em 1...
sheet.cell(row=2, column=1).value = "Daxus"
sheet.cell(2, 2).value = "Automação no Python"

# Apagando dados (podem ser por rótulo ou coordenadas)
sheet["A2"].value = None
sheet.cell(2, 2).value = None

wb.save('comandos_basicos.xlsx')

