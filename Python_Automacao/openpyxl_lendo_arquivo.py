from openpyxl import load_workbook

wb = load_workbook("Lendo Excel.xlsx")

# Verificando as abas existentes
print(wb.sheetnames)
print(wb.sheetnames[0])
print(wb.sheetnames[1])
print(wb.sheetnames[2])

# Acessando as abas
sheet_nomes = wb["Nomes"]
sheet_semanas = wb["Dias da semana"]
sheet_doces = wb["Doces"]

print(sheet_nomes["A2"].value)
print(sheet_semanas.cell(8, 1).value)

# Atualizando o valor de uma célula
sheet_doces["A6"].value = "Bomba doce de leite"

# Salvando o arquivo com as alterações
wb.save("Lendo Excel.xlsx")