from openpyxl import load_workbook

# Carregando o Arquivo
wb = load_workbook("Modificando estrutura.xlsx")

# Verificando as abas existentes no arquivo
print(wb.sheetnames)

# Selecionando a aba existente
sheet = wb["Vendas"]

# Renomeando uma aba
sheet.title = "Vendas 2026"

# Criando uma nova aba
wb.create_sheet("Vendas 2027")

# Mesclando células
sheet.merge_cells("A1:D1")

# Testando com a aba Vendas 2027
sheet_2 = wb["Vendas 2027"]
sheet_2.merge_cells("B1:F1")

# Retirando a mesclagem de céluas
sheet.unmerge_cells("A1:D1")
sheet_2.unmerge_cells("B1:F1")

# Inserindo novas linhas
# Linha 4 será reposicionada abaixo
sheet.insert_rows(3)

# Inserindo novas colunas
sheet.insert_cols(2)

# Deletendo linhas e colunas
sheet.delete_rows(3)
sheet.delete_cols(2)

# Salvando o arquivo
wb.save("Modificando estrutura.xlsx")