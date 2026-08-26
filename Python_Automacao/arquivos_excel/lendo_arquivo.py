from openpyxl import load_workbook

wb = load_workbook("dados/Lendo Excel.xlsx")

# Verificando as abas existentes
print(wb.sheetnames)
print(wb.sheetnames[0])
print(wb.sheetnames[1])
print(wb.sheetnames[2])

# Acessando as abas
sheet_nomes = wb["Nomes"]
sheet_semanas = wb["Dias da semana"]
sheet_doces = wb["Doces"]

print(sheet_nomes["A2"].value)
print(sheet_semanas.cell(8, 1).value)

# Atualizando o valor de uma célula
sheet_doces["A6"].value = "Bomba doce de leite"

# Descobrindo total de linhas e e colunas
print(f"Guia/ Aba Nomes, possui {sheet_nomes.max_row} linhas.")
print(f"Guia/ Aba Nomes, possui {sheet_nomes.max_column} colunas.")

print(type(sheet_nomes.max_row))

# Percorrendo os dados de uma aba

# Iniciar o range na linha 2 para não pegar o cabeçalho
for linha in range(2, sheet_nomes.max_row + 1): # incluir o +1 pois o último número no range é exclusivo
    print(sheet_nomes.cell(linha, 1).value)

# Melhorando o for: pegando as linhas e as colunas
for linha in range(2, sheet_nomes.max_row + 1): # incluir o +1 pois o último número no range é exclusivo
    for coluna in range(1, sheet_nomes.max_column + 1):
        print(sheet_nomes.cell(linha, coluna).value)

# Criando uma lista com as informações:
listas = []
for linha in range(2, sheet_nomes.max_row + 1): # incluir o +1 pois o último número no range é exclusivo
    for coluna in range(1, sheet_nomes.max_column + 1):
        listas.append(sheet_nomes.cell(linha, coluna).value)
listas
# print(type(listas))

# Criando um dicionário:
chaves = []
valores = []
dicionario = {}
for linha in range(2, sheet_nomes.max_row + 1): # incluir o +1 pois o último número no range é exclusivo
    for coluna in range(1, sheet_nomes.max_column + 1):
        chaves = sheet_nomes.cell(linha, 1).value
        valores = sheet_nomes.cell(linha, 2).value
        dicionario[chaves] = valores
dicionario
# print(type(dicionario))


# Salvando o arquivo com as alterações
wb.save("dados/Lendo Excel.xlsx")