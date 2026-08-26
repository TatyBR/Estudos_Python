from openpyxl import load_workbook
# FORMULAE: lista de fórmulas Excel
from openpyxl.utils import FORMULAE

wb = load_workbook("dados/Fórmulas.xlsx")

# Acessando uma aba/ Só quando existe uma única aba
# Em vez de: sheet = wb["Vendas"]
sheet = wb.active

# Criando o cabeçalho da coluna E
sheet["E1"] = "Valor Total"

# Testa forma, o cálculo é feito mas o arquivo aberto no Excel só exibe o valor e não a Fórmula
# sheet["E2"] = sheet["C2"].value * sheet["D2"].value

# Fazendo desta forma abaixo, aparece a Fórmula utilizada!!
sheet["E2"] = "=C2*D2"

# Preenchendo demais linhas...
for linha in range(3, 12):
    sheet[f"E{linha}"] = f"=C{linha}*D{linha}"

# Preenchendo 
sheet["F1"] = "Valor Total2"

for linha in range(2, sheet.max_row):
    sheet[f"F{linha}"] = f"=C{linha}*D{linha}"

# Verificando todas as fórmulas diponíveis
# FORMULAE
# ou
# for i in FORMULAE:
    # print(i)

# Utilizando uma fórmula pronta:
# Tem que ser usado o nome em Inglês (consultar Formulae)
sheet["E12"] = "=SUM(E2:E11)"

# Incluindo a soma total na coluna F:
linha_final = sheet.max_row
sheet[f"F{linha_final}"] = f"=SUM(F2:F{linha_final})"

# Salvando o arquivo
wb.save("dados/Fórmulas.xlsx")