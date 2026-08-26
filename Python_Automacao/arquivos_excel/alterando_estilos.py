# Documentação ref. a estilos: https://openpyxl.readthedocs.io/en/stable/styles.html

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

wb = load_workbook("dados/Estilos.xlsx")

# Ativando a aba
sheet = wb["Estilos"]

# Alterando a cor da fonte
sheet["A1"].font = Font(color="000000FF")

# Preenchimento da célula
sheet["A2"].fill = PatternFill("darkDown","00FFFF00")

# Alterando a fonte
sheet["A3"].font = Font(name="Arial")

# Aplicando Negrito e Itálico
sheet["A4"].font = Font(bold=True)
sheet["A5"].font = Font(italic=True)

# Alterando o alinhamento da célula
sheet["A6"].alignment = Alignment(horizontal="center", vertical="center")

# Salvando o arquivo
wb.save("dados/Estilos.xlsx")